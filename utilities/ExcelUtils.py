import openpyxl
class ExcelReadWriteCount:
    def rowCount(path,sheetname):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.get_sheet_by_name(sheetname)
        return(sheet.max_row)

    def colCount(path,sheetname):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.get_sheet_by_name(sheetname)
        return(sheet.max_column)

    def readExcelData(path,sheetname,rownum,colnum):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook.get_sheet_by_name(sheetname)
        return(sheet.cell(rownum,colnum).value)

    def writeExcelData(path,sheetname,rownum,colnum,data):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name(sheetname)
        sheet.cell(rownum,colnum).value=data
        workbook.save(path)